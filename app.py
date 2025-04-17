from fastapi import FastAPI, Request, Response, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import pandas as pd
import io
import math
import datetime
import os
import uvicorn
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from pydantic import BaseModel
from typing import List, Optional
import shutil

app = FastAPI(title="Cable Sizing Calculator")

# Create necessary directories
os.makedirs("templates", exist_ok=True)
os.makedirs("static", exist_ok=True)

# Copy logo file to static directory if it doesn't exist yet
if os.path.exists("vector_logo1.png") and not os.path.exists("static/vector_logo1.png"):
    shutil.copy("vector_logo1.png", "static/vector_logo1.png")

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Templates
templates = Jinja2Templates(directory="templates")

# Default values
DEFAULT_VOLTAGE = 100

# Cable data from example
DEFAULT_CABLE_DATA = [
    [15, 25],    # Cable 1
    [15, 25],    # Cable 2
    [26, 6],     # Cable 3
    [25, 25],    # Cable 4
    [20, 25],    # Cable 5
    [20, 25],    # Cable 6
    [38, 25],    # Cable 7
    [37, 25],    # Cable 8
    [26, 25],    # Cable 9
    [39, 25],    # Cable 10
    [0, 0],      # Cable 11
    [0, 0]       # Cable 12
]

# Define data models
class Cable(BaseModel):
    cableType: float
    cableLength: float
    powerTapping: float
    voltageDrop: Optional[float] = None
    voltageAtSpeaker: Optional[float] = None
    splReduction: Optional[float] = None
    examineStatus: Optional[str] = None

class CalculateRequest(BaseModel):
    cables: List[Cable]

class ExcelRequest(BaseModel):
    cables: List[Cable]
    numSpeakers: int
    defaultCableType: float

@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/calculate")
def calculate(calc_request: CalculateRequest):
    results = []
    previous_voltage = DEFAULT_VOLTAGE
    
    for cable in calc_request.cables:
        cable_type = cable.cableType
        cable_length = cable.cableLength
        power_tapping = cable.powerTapping
        
        # Calculate voltage drop using the formula
        voltage_drop = 1.68 * math.pow(10, -8) * cable_length * power_tapping / (100 * cable_type * math.pow(10, -6))
        
        # Calculate voltage at speaker
        voltage_at_speaker = previous_voltage - voltage_drop
        
        # Calculate SPL reduction
        spl_reduction = 20 * math.log10(voltage_at_speaker / 100) if voltage_at_speaker > 0 else -99.999
        
        # Determine examine status
        examine_status = 'Normal' if spl_reduction > -2 else 'Abnormal'
        
        # Create result object
        result = {
            'voltageDrop': round(voltage_drop, 3),
            'voltageAtSpeaker': round(voltage_at_speaker, 3),
            'splReduction': round(spl_reduction, 3) if spl_reduction != -99.999 else -99.999,
            'examineStatus': examine_status
        }
        
        results.append(result)
        
        # Update previous voltage for next calculation
        previous_voltage = voltage_at_speaker
    
    return results

@app.post("/generate_excel")
def generate_excel(excel_request: ExcelRequest):
    # Filter cables with power tapping > 0
    filtered_cables = [c for c in excel_request.cables if c.powerTapping > 0]
    
    # Import font for styling
    from openpyxl.styles import Font, PatternFill
    
    # Create fill colors for Normal/Abnormal status
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Copy the template file to a new file to preserve all images and formatting
    template_path = "cable2.xlsx"
    output_path = "cable_sizing_output.xlsx"
    
    # Direct file copy (preserves all content including images)
    if os.path.exists(template_path):
        shutil.copy2(template_path, output_path)
        
        # Now modify only the cells we need to change
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Get the first row index with data (skip headers)
        data_start_row = 12
        
        # Create font style with size 12
        font_12 = Font(name="Arial", size=12)
        
        # Fill data rows
        for i, cable in enumerate(filtered_cables):
            row_idx = data_start_row + i
            
            # Write Cable number (1, 2, 3, etc.)
            cell = ws.cell(row=row_idx, column=1, value=i + 1)
            cell.font = font_12
            
            # Write user input values
            cell = ws.cell(row=row_idx, column=2, value=cable.cableType)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=3, value=cable.cableLength)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=4, value=cable.powerTapping)
            cell.font = font_12
            
            # Write calculated values with proper formatting - use values from UI
            cell = ws.cell(row=row_idx, column=5, value=cable.voltageDrop)
            cell.number_format = '0.000'
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=6, value=cable.voltageAtSpeaker)
            cell.number_format = '0.000'
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=7, value=cable.splReduction)
            cell.number_format = '0.000'
            cell.font = font_12
            
            # Write examine status (Normal or Abnormal) with color formatting
            cell = ws.cell(row=row_idx, column=8, value=cable.examineStatus)
            cell.font = font_12
            
            # Apply conditional fill colors
            if cable.examineStatus == "Normal":
                cell.fill = green_fill
            else:
                cell.fill = red_fill
        
        # Save the modified file
        wb.save(output_path)
        
        # Return the file 
        return FileResponse(
            path=output_path,
            filename="cable_sizing.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=cable_sizing.xlsx",
                "X-Content-Type-Options": "nosniff",
                "Access-Control-Allow-Origin": "*"
            }
        )
    else:
        # Template doesn't exist, create a new file
        wb = Workbook()
        ws = wb.active
        ws.title = "Cable Sizing"
        
        # Create font style with size 12
        font_12 = Font(name="Arial", size=12)
        
        # Fill data rows
        for i, cable in enumerate(filtered_cables):
            row_idx = i + 1
            
            # Write all cells with font size 12
            cell = ws.cell(row=row_idx, column=1, value=i + 1)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=2, value=cable.cableType)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=3, value=cable.cableLength)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=4, value=cable.powerTapping)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=5, value=cable.voltageDrop)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=6, value=cable.voltageAtSpeaker)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=7, value=cable.splReduction)
            cell.font = font_12
            
            cell = ws.cell(row=row_idx, column=8, value=cable.examineStatus)
            cell.font = font_12
            
            # Apply conditional fill colors
            if cable.examineStatus == "Normal":
                cell.fill = green_fill
            else:
                cell.fill = red_fill
        
        # Save to a file
        wb.save(output_path)
        
        return FileResponse(
            path=output_path,
            filename="cable_sizing.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

@app.post("/upload_excel")
async def upload_excel(file: UploadFile = File(...)):
    try:
        # Read the uploaded Excel file
        contents = await file.read()
        
        # Load workbook from uploaded file
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Define where data starts in the Excel sheet (row 12 from the image)
        data_start_row = 12
        
        # Extract data from Excel
        cables = []
        row_idx = data_start_row
        
        # Continue reading rows until we find an empty row or reach a reasonable limit
        while row_idx < 100:  # Set a reasonable limit to avoid infinite loops
            # Read cable data from each row
            cable_type = ws.cell(row=row_idx, column=2).value
            cable_length = ws.cell(row=row_idx, column=3).value
            power_tapping = ws.cell(row=row_idx, column=4).value
            
            # Skip empty rows or rows without essential data
            if not power_tapping or not cable_type or not cable_length:
                row_idx += 1
                continue
                
            # Convert values to appropriate types
            try:
                cable_type = float(cable_type)
                cable_length = float(cable_length)
                power_tapping = float(power_tapping)
            except (TypeError, ValueError):
                # Skip rows with non-numeric values
                row_idx += 1
                continue
                
            # Read calculated values if they exist
            voltage_drop = ws.cell(row=row_idx, column=5).value
            voltage_at_speaker = ws.cell(row=row_idx, column=6).value
            spl_reduction = ws.cell(row=row_idx, column=7).value
            examine_status = ws.cell(row=row_idx, column=8).value
            
            # Create cable object
            cable = {
                "cableType": cable_type,
                "cableLength": cable_length,
                "powerTapping": power_tapping,
                "voltageDrop": float(voltage_drop) if voltage_drop is not None else None,
                "voltageAtSpeaker": float(voltage_at_speaker) if voltage_at_speaker is not None else None,
                "splReduction": float(spl_reduction) if spl_reduction is not None else None,
                "examineStatus": examine_status if examine_status is not None else None
            }
            
            cables.append(cable)
            row_idx += 1
            
            # Stop if we've gathered a reasonable amount of data
            if len(cables) >= 20:
                break
        
        # Get the number of speakers (number of rows with data)
        num_speakers = len(cables)
        
        # Get the default cable type (from the first row)
        default_cable_type = cables[0]["cableType"] if cables else 2.5
        
        # Return the extracted data
        return {
            "success": True,
            "cables": cables,
            "numSpeakers": num_speakers,
            "defaultCableType": default_cable_type
        }
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )

if __name__ == '__main__':
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True) 